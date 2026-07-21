"""Compare simulated workload indicators against external NASA-TLX observations.

This module is part of the BVI ACT-R navigation simulation workflow.
"""

import argparse
import contextlib
import datetime
import io
import json
import math
import multiprocessing as mp
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent
if not (ROOT / "bvi_sa").exists():
    ROOT = ROOT.parent
XLSX = ROOT.parent / "NASA-TLX统计表(1).xlsx"
CITY = {"P01": "杭州", "P02": "江阴", "P03": "上海"}
OUT = ROOT / "calib_out_v2"
SEEDS = [20260705 + i for i in range(5)]

DIMS = ["心理需求", "身体需求", "时间压力", "自我表现", "努力程度", "挫败感"]
TLX_HIGH = 60.0
IW_HIGH = 6.0


def parse_score(v):
    """Parse a questionnaire or TLX score cell."""
    if v is None or str(v).strip() == "":
        return float("nan")
    if isinstance(v, (datetime.datetime, datetime.date)):
        return (v.month + v.day) / 2.0
    s = str(v).strip()
    m = re.fullmatch(r"(\d+(?:\.\d+)?)\s*[-–~]\s*(\d+(?:\.\d+)?)", s)
    if m:
        return (float(m.group(1)) + float(m.group(2))) / 2.0
    if re.fullmatch(r"\d+(?:\.\d+)?", s):
        return float(s)
    return float("nan")


def _mk_seg(city, trial, typ, familiar, scores):
    """Create one external-validity segment record."""
    comp = [s for s in scores.values() if not math.isnan(s)]
    return {
        "city": city,
        "trial": trial,
        "type": typ,
        "familiar": familiar,
        "scores": scores,
        "composite": sum(comp) / len(comp) if comp else float("nan"),
    }


def load_all():
    """Load all external-validity observation segments."""
    import openpyxl

    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    starts = [i for i, r in enumerate(rows) if str(r[0] or "").strip() == "被试编号"]
    segs = []
    for bi, st in enumerate(starts):
        end = starts[bi + 1] if bi + 1 < len(starts) else len(rows)
        block = rows[st:end]

        def row_of(key):
            """Find the row matching a questionnaire key."""
            for r in block:
                if str(r[0] or "").startswith(key):
                    return r
            return None

        r_pid, r_fam = row_of("被试编号"), row_of("路线熟悉度")
        r_trial, r_type = row_of("试次编号"), row_of("分段类型")
        dim_rows = {d: row_of(d) for d in DIMS}
        assert all(dim_rows.values()), f"块{bi}缺维度行"
        for c in range(1, len(r_pid)):
            if not r_trial[c] or not (r_type[c] and str(r_type[c]).strip()):
                continue
            pid = str(r_pid[c]).strip()
            fam = "Unfamiliar" not in str(r_fam[c])
            typ = str(r_type[c]).replace("\n", "").strip()
            scores = {d: parse_score(dim_rows[d][c]) for d in DIMS}
            segs.append(_mk_seg(CITY.get(pid, pid), str(r_trial[c]), typ, fam, scores))
    return segs


def tlx_indicators(segs):
    """Compute TLX-derived workload indicators."""
    comp = [s["composite"] for s in segs]
    junc = [s["composite"] for s in segs if "路口" in s["type"]]
    walk = [s["composite"] for s in segs if "路口" not in s["type"]]
    out = {
        "n_segments": len(segs),
        "mean": sum(comp) / len(comp) if comp else float("nan"),
        "peak": max(comp) if comp else float("nan"),
        "high_ratio": (
            sum(1 for c in comp if c >= TLX_HIGH) / len(comp) if comp else float("nan")
        ),
        "junction_mean": sum(junc) / len(junc) if junc else float("nan"),
        "nonjunction_mean": sum(walk) / len(walk) if walk else float("nan"),
    }
    out["junction_minus_non"] = (
        out["junction_mean"] - out["nonjunction_mean"]
        if junc and walk
        else float("nan")
    )
    return out


def _extract_load(
    sim_log,
    event_log,
    profile,
    steps,
    start_node,
    goal_node,
    current_position,
    max_steps,
    graph=None,
    initial_production_utilities=None,
    report_dir=None,
):
    """Handle extract load behavior."""
    n = len(sim_log)
    iw = [float(r.get("actr_iw_total", 0.0)) for r in sim_log]
    crossing = [bool(r.get("crossing_active")) for r in sim_log]
    iw_sorted = sorted(iw)
    p95 = iw_sorted[min(n - 1, int(math.ceil(0.95 * n)) - 1)]
    j = [iw[i] for i in range(n) if crossing[i]]
    w = [iw[i] for i in range(n) if not crossing[i]]
    return {
        "iw_mean": sum(iw) / n,
        "iw_p95": p95,
        "iw_high_ratio": sum(1 for v in iw if v >= IW_HIGH) / n,
        "iw_junction_mean": sum(j) / len(j) if j else float("nan"),
        "iw_nonjunction_mean": sum(w) / len(w) if w else float("nan"),
        "_steps": n,
    }


_SIM = None


def _init():
    """Initialize worker-process imports and paths."""
    global _SIM
    with contextlib.redirect_stdout(io.StringIO()):
        sys.path.insert(0, str(ROOT))
        from bvi_sa import simulation as s
    s.generate_report = _extract_load
    _SIM = s


def _run(task):
    """Execute one simulation task and return metrics."""
    seed, fam = task
    import random

    try:
        random.seed(seed)
        import numpy as np

        np.random.seed(seed % (2**32 - 1))
        with contextlib.redirect_stdout(io.StringIO()):
            m = _SIM.run_simulation(familiarity_level=fam)
        m["_seed"], m["_fam"] = seed, fam
        return m, None
    except Exception as e:
        return None, f"{type(e).__name__}: {e}"


KEYS = ["iw_mean", "iw_p95", "iw_high_ratio", "iw_junction_mean", "iw_nonjunction_mean"]


def agg(results):
    """Aggregate simulation results into summary metrics."""
    m = {k: sum(r[k] for r in results) / len(results) for k in KEYS}
    m["iw_junction_minus_non"] = m["iw_junction_mean"] - m["iw_nonjunction_mean"]
    m["per_seed"] = {k: [r[k] for r in results] for k in KEYS}
    return m


def main():
    """Run the script entry point."""
    ap = argparse.ArgumentParser()
    ap.add_argument("--parse-only", action="store_true")
    args = ap.parse_args()

    segs = load_all()
    fam = [s for s in segs if s["familiar"]]
    unf = [s for s in segs if not s["familiar"]]
    print(
        f"TLX 共 {len(segs)} 段（杭州 {sum(s['city']=='杭州' for s in segs)}，"
        f"江阴 {sum(s['city']=='江阴' for s in segs)}，上海 {sum(s['city']=='上海' for s in segs)}；"
        f"熟悉 {len(fam)}，陌生 {len(unf)}）"
    )
    for s in segs:
        print(
            f"  {s['city']} {s['trial']:<4}{s['type']:<7}"
            f"{'F' if s['familiar'] else 'U'}  综合={s['composite']:.1f}"
        )

    t = {
        "familiar": tlx_indicators(fam),
        "unfamiliar": tlx_indicators(unf),
        "all": tlx_indicators(segs),
    }
    for city in ("杭州", "江阴", "上海"):
        t[f"familiar_{city}"] = tlx_indicators([s for s in fam if s["city"] == city])
    print(
        "\nTLX 指标（熟悉段合并，主口径）:",
        {k: round(v, 3) for k, v in t["familiar"].items()},
    )
    print(
        "TLX 指标（陌生段合并）:", {k: round(v, 3) for k, v in t["unfamiliar"].items()}
    )
    if args.parse_only:
        return

    ctx = mp.get_context("spawn")
    tasks = [(sd, 1) for sd in SEEDS] + [(sd, 0) for sd in SEEDS]
    with ctx.Pool(processes=10, initializer=_init) as pool:
        results = [m for m, err in pool.map(_run, tasks) if m]
    model_f = agg([r for r in results if r["_fam"] == 1])
    model_u = agg([r for r in results if r["_fam"] == 0])
    print("\n模型（熟悉，5种子）:", {k: round(model_f[k], 3) for k in KEYS})
    print("模型（陌生，5种子）:", {k: round(model_u[k], 3) for k in KEYS})

    comp = {
        "mean": {"tlx": t["familiar"]["mean"] / 100, "model": model_f["iw_mean"] / 10},
        "peak": {"tlx": t["familiar"]["peak"] / 100, "model": model_f["iw_p95"] / 10},
        "high_ratio": {
            "tlx": t["familiar"]["high_ratio"],
            "model": model_f["iw_high_ratio"],
        },
        "junction_minus_non": {
            "tlx": t["familiar"]["junction_minus_non"] / 100,
            "model": model_f["iw_junction_minus_non"] / 10,
        },
        "unfamiliar_minus_familiar_mean": {
            "tlx": (t["unfamiliar"]["mean"] - t["familiar"]["mean"]) / 100,
            "model": (model_u["iw_mean"] - model_f["iw_mean"]) / 10,
        },
    }
    print("\n归一化对比:")
    for k, v in comp.items():
        print(f"  {k}: TLX={v['tlx']:.3f}  模型={v['model']:.3f}")

    out = {
        "tlx_segments": segs,
        "tlx": t,
        "model_familiar": model_f,
        "model_unfamiliar": model_u,
        "normalized_comparison": comp,
        "notes": [
            "TLX 为分段回溯评分，与仿真逐步时序不可对齐，仅做聚合对比",
            "主对比口径为三城熟悉段合并，与仿真 familiarity_level=1 对应",
            "数据源为 0717 修订版 NASA-TLX统计表.xlsx（三城三被试，信度低评分已纠正）",
            "高负荷阈值：TLX 60/100，模型 IW 6/10（均为量程 60%）",
        ],
    }
    with open(OUT / "tlx_external_validity.json", "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print(f"\n已写入 {OUT / 'tlx_external_validity.json'}")


if __name__ == "__main__":
    main()
